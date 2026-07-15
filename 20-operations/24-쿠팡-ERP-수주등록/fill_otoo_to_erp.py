# -*- coding: utf-8 -*-
"""
쿠팡 로켓/프레시 OTOO 발주 파일 -> 사내 ERP 위탁수주 일괄입력 양식 자동 채움.

흐름:
  1. 지정된 날짜(기본: 실행일 기준 전일)의 발주서 폴더에서 OTOO 파일(들)을 찾는다.
  2. 파일 내용을 읽어 ERP 양식 컬럼(A~BS)에 맞게 변환한다.
  3. ERP 위탁수주 파일(로켓/프레시)에 그 날짜 이름의 새 시트를 만들어 채운다.

안전장치:
  - 기본은 --write 없이 dry-run: 무엇을 할지만 보여주고 실제로는 아무것도 쓰지 않음.
  - 대상 시트가 이미 있으면 절대 덮어쓰지 않고 건너뜀.
  - 대상 파일이 현재 엑셀에서 열려 있으면(잠금파일 ~$ 존재) 쓰기를 거부.
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, timedelta

import openpyxl
from openpyxl.utils import column_index_from_string

BASE = Path(r"G:\공유 드라이브\HQ_Team_SCM\SCM_공용\SCM_물류운영")
SRC_ROOT = BASE / "발주서" / "직매입"
DST_ROOT = BASE / "ERP 데이터" / "위탁수주"

# 고정값 (기존에 완료된 시트들과 대조하여 확인됨 — 로켓/프레시 공통)
FIXED_VALUES = {
    "A": "SOT06",
    "B": "1000",
    "C": "00",
    "D": "00",
    "K": "egnis",
    "L": "1000",
    "M": "1000",
    "N": "1000",
    "O": "4660",
    "Q": "KRW",
    "R": 1,
    "S": "EA",
    "T": "1",
    "W": "1",
    "AK": "B",
    "AL": "11",
    "AM": 1,
    "AW": "쿠팡",
    "AX": "010-1111-1111",
    "AY": "010-1111-1111",
    "AZ": "449823",
    "BA": "경기도 용인시 처인구 양지면 양지리 120-12",
    "BD": "20",
}

CHANNELS = {
    "로켓": {
        "src_folder": "쿠팡 로켓",
        "dst_file_word": "로켓",
    },
    "프레시": {
        "src_folder": "쿠팡 프레시",
        "dst_file_word": "프레시",
    },
}


def prev_date(base=None):
    base = base or datetime.now()
    return (base - timedelta(days=1)).date()


def rocket_folder_name(d):
    return f"{d.year % 100:02d}{d.month:02d}{d.day:02d}"


def fresh_folder_name(d):
    return f"{d.year % 100:02d}.{d.month:02d}.{d.day:02d}"


def rocket_sheet_name(d):
    return f"{d.month}{d.day:02d}"


def fresh_sheet_name(d):
    return f"{d.month:02d}{d.day:02d}"


def find_source_folder(channel, d):
    year_dir = SRC_ROOT / CHANNELS[channel]["src_folder"] / str(d.year)
    folder_name = rocket_folder_name(d) if channel == "로켓" else fresh_folder_name(d)
    folder = year_dir / folder_name
    return folder if folder.is_dir() else None


def find_otoo_files(folder):
    files = []
    for p in folder.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if "OTOO" in p.name and p.suffix.lower() in (".xlsx", ".xls"):
            files.append(p)
    # 본 파일(추가/재고부족이 안 붙은 것) 먼저, 그 다음 부가 파일
    files.sort(key=lambda p: (("추가" in p.name) or ("재고부족" in p.name), p.name))
    return files


def read_otoo_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = []
    for r in range(2, ws.max_row + 1):
        item_code = ws.cell(row=r, column=21).value  # 품목코드
        if item_code is None:
            continue
        rows.append({
            "거래처코드": ws.cell(row=r, column=4).value,
            "품목코드": item_code,
            "품목명": ws.cell(row=r, column=22).value,
            "주문요청수량": ws.cell(row=r, column=23).value,
            "주문금액": ws.cell(row=r, column=25).value,
        })
    return rows


def find_dest_file(channel, d):
    month_folder = f"{d.year % 100:02d}.{d.month:02d}"
    folder = DST_ROOT / month_folder / "직매입"
    if not folder.is_dir():
        return None
    word = CHANNELS[channel]["dst_file_word"]
    suffix = f"({d.month}월 {word}).xlsx"
    candidates = [
        p for p in folder.iterdir()
        if p.is_file() and not p.name.startswith("~$") and p.name.endswith(suffix)
    ]
    return candidates[0] if candidates else None


def is_locked(path):
    return (path.parent / f"~${path.name}").exists()


def build_rows(otoo_rows, d):
    date_int = int(f"{d.year}{d.month:02d}{d.day:02d}")
    out = []
    for r in otoo_rows:
        row = dict(FIXED_VALUES)
        row["G"] = r["거래처코드"]
        row["P"] = date_int
        row["U"] = r["품목코드"]
        row["V"] = r["품목명"]
        row["Y"] = r["주문요청수량"]
        row["AJ"] = r["주문금액"]
        row["AN"] = date_int
        row["BE"] = date_int
        out.append(row)
    return out


def process_channel(channel, d, write=False):
    result = {"channel": channel, "date": str(d)}
    src_folder = find_source_folder(channel, d)
    if src_folder is None:
        result["status"] = "no_source_folder"
        return result

    otoo_files = find_otoo_files(src_folder)
    if not otoo_files:
        result["status"] = "no_otoo_file"
        result["src_folder"] = str(src_folder)
        return result

    all_rows = []
    for f in otoo_files:
        all_rows.extend(read_otoo_rows(f))

    dest_file = find_dest_file(channel, d)
    if dest_file is None:
        result["status"] = "no_dest_file"
        return result

    sheet_name = rocket_sheet_name(d) if channel == "로켓" else fresh_sheet_name(d)
    built_rows = build_rows(all_rows, d)

    result.update({
        "status": "ready",
        "src_folder": str(src_folder),
        "otoo_files": [f.name for f in otoo_files],
        "dest_file": str(dest_file),
        "sheet_name": sheet_name,
        "row_count": len(built_rows),
        "preview_first": built_rows[0] if built_rows else None,
        "preview_last": built_rows[-1] if built_rows else None,
    })

    if not write:
        result["write"] = "skipped (dry-run)"
        return result

    if is_locked(dest_file):
        result["write"] = "ABORTED: dest file is open in Excel (lock file present)"
        return result

    wb = openpyxl.load_workbook(dest_file)
    if sheet_name in wb.sheetnames:
        result["write"] = f"ABORTED: sheet '{sheet_name}' already exists"
        return result

    template = wb[wb.sheetnames[0]]
    ws = wb.create_sheet(sheet_name)
    for r in range(1, 4):
        for c in range(1, template.max_column + 1):
            ws.cell(row=r, column=c).value = template.cell(row=r, column=c).value

    for i, row in enumerate(built_rows, start=4):
        for col_letter, value in row.items():
            ws.cell(row=i, column=column_index_from_string(col_letter)).value = value

    wb.save(dest_file)
    result["write"] = "done"
    return result


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", help="YYYY-MM-DD (기본: 실행일 전일)")
    ap.add_argument("--write", action="store_true", help="실제로 파일에 씀 (기본은 dry-run)")
    args = ap.parse_args()

    d = datetime.strptime(args.date, "%Y-%m-%d").date() if args.date else prev_date()

    results = [process_channel(ch, d, write=args.write) for ch in CHANNELS]
    print(json.dumps(results, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
